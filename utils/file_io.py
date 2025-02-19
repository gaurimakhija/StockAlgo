import json
import csv
from pathlib import Path
import openpyxl as excel

class CSVFileHandler:
    
    def __init__(self, filepath: Path):
        self.filepath = Path(filepath).resolve()

    def read_csv(self) -> list:
        """Reads a CSV file and returns its contents as a list of rows."""
        data = []
        with open(self.filepath, 'r', newline='') as file:
            reader = csv.reader(file)
            for row in reader:
                data.append(row)
        return data

    def write_csv(self, data: list):
        """Writes a list of rows to a CSV file."""
        with open(self.filepath, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(data)  # More efficient than looping row by row


class JSONFileHandler:
    
    def __init__(self, filepath: Path):
        self.filepath = Path(filepath).resolve()

    def read_json(self) -> dict:
        """Reads a JSON file and returns its contents as a dictionary."""
        with open(self.filepath, 'r') as file:
            return json.load(file)

    def write_json(self, data: dict):
        """Writes a dictionary to a JSON file."""
        with open(self.filepath, 'w') as file:
            json.dump(data, file, indent=4)  # Indent for better readability


class ExcelFileHandler:
    
    def __init__(self, filepath: Path):
        self.filepath = Path(filepath).resolve()

    def read_excel(self) -> dict:
        """Reads an Excel file and returns a dictionary with sheet names as keys."""
        wb = excel.load_workbook(self.filepath)
        data = {}
        for sheet in wb.sheetnames:
            data[sheet] = [row for row in wb[sheet].iter_rows(values_only=True)]
        return data

    def write_excel(self, data: dict):
        """Writes data to an Excel file, creating sheets as needed."""
        wb = excel.Workbook()
        default_sheet = wb.active  # Default created sheet
        if len(data) == 1:
            sheet_name = next(iter(data.keys()))  # Use the only sheet's name
            ws = default_sheet
            ws.title = sheet_name
        else:
            wb.remove(default_sheet)  # Remove default sheet if multiple sheets exist
        
        for sheet, rows in data.items():
            ws = wb.create_sheet(sheet) if len(data) > 1 else default_sheet
            for row in rows:
                ws.append(row)
        wb.save(self.filepath)
    